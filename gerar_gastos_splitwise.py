# -*- coding: utf-8 -*-
"""Gastos compartilhados + pessoais (Splitwise) — Google Sheets. Gera gastos-compartilhados.xlsx."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from pathlib import Path

MEMBROS = ["Eduardo", "Karina", "Bruno", "Gabriele", "Henry"]
N_ROWS = 120
LAST = N_ROWS + 1
COL_PART = {m: get_column_letter(8 + i) for i, m in enumerate(MEMBROS)}  # H..L

wb = Workbook()
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(color="FFFFFF", bold=True, size=11)
title_font = Font(bold=True, size=16, color="1F4E79")
neg_fill = PatternFill("solid", fgColor="FFC7CE")
pos_fill = PatternFill("solid", fgColor="C6EFCE")


def style_header(ws, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(1, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def autosize(ws, max_width=48):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col:
            if cell.value:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


# ---------- Instrucoes ----------
ws = wb.active
ws.title = "Instrucoes"
ws["A1"] = "Gastos da viagem · Orlando 10–24 nov 2026"
ws["A1"].font = title_font
lines = [
    "",
    "Membros: " + ", ".join(MEMBROS),
    "",
    "Como usar no Google Sheets",
    "1) Abra a aba Gastos e marque TRUE/FALSE (ou caixa de seleção) em quem divide cada gasto compartilhado.",
    "2) Tipo = Compartilhado → entra no saldo entre vocês. Tipo = Pessoal → só no seu custo da viagem.",
    "3) A aba Resumo calcula saldo (quem deve a quem) e custo total da viagem por pessoa.",
    "4) Registre Pix/Zelle na aba Acertos.",
    "",
    "Incluir Henry (planilha já existente no Drive)",
    "• Inserir coluna Henry após Gabriele na aba Gastos (participa?).",
    "• Atualizar fórmula Div: =IF(D2=\"Pessoal\",1,COUNTIF(H2:L2,TRUE))",
    "• Copiar linha de Gabriele no Resumo para Henry e trocar coluna H→L nas fórmulas.",
    "• Ou: importar este .xlsx de novo (substituir abas Gastos, Resumo, Membros).",
]
for i, text in enumerate(lines, start=2):
    ws.cell(i, 1, text)
    if text.startswith("Como") or text.startswith("Incluir"):
        ws.cell(i, 1).font = Font(bold=True, color="1F4E79")
ws.column_dimensions["A"].width = 95

# ---------- Membros ----------
ws = wb.create_sheet("Membros")
ws.append(["Membro", "Ativo"])
style_header(ws, 2)
for m in MEMBROS:
    ws.append([m, True])
autosize(ws)

# ---------- Gastos ----------
ws = wb.create_sheet("Gastos")
headers = ["Data", "Descricao", "Categoria", "Tipo", "Moeda", "Valor", "Pagou"]
headers += MEMBROS + ["Div", "Obs"]
ws.append(headers)
style_header(ws, len(headers))

dv_tipo = DataValidation(type="list", formula1='"Compartilhado,Pessoal"', allow_blank=True)
dv_moeda = DataValidation(type="list", formula1='"USD,BRL"', allow_blank=True)
dv_pagou = DataValidation(type="list", formula1='"%s"' % ",".join(MEMBROS), allow_blank=True)
dv_bool = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
for dv in (dv_tipo, dv_moeda, dv_pagou, dv_bool):
    ws.add_data_validation(dv)

part_cols = [get_column_letter(8 + i) for i in range(len(MEMBROS))]
part_range = ":".join((part_cols[0], part_cols[-1]))
div_col = get_column_letter(8 + len(MEMBROS))

for r in range(2, LAST + 1):
    ws.cell(r, 4, "Compartilhado")
    ws.cell(r, 5, "USD")
    for c in range(8, 8 + len(MEMBROS)):
        ws.cell(r, c, False)
    ws.cell(r, len(headers) - 1, f'=IF(D{r}="Pessoal",1,COUNTIF({part_cols[0]}{r}:{part_cols[-1]}{r},TRUE))')
    dv_tipo.add(f"D{r}")
    dv_moeda.add(f"E{r}")
    dv_pagou.add(f"G{r}")
    for c in range(8, 8 + len(MEMBROS)):
        dv_bool.add(f"{get_column_letter(c)}{r}")

ws.freeze_panes = "A2"
autosize(ws)

# ---------- Resumo ----------
ws = wb.create_sheet("Resumo")
ws["A1"] = "Resumo por membro"
ws["A1"].font = title_font
ws.append([])
headers = ["Membro", "Total pagou", "Cota comum", "Gastos pessoais", "Custo da viagem", "Saldo Splitwise"]
ws.append(headers)
style_header(ws, len(headers))

for i, membro in enumerate(MEMBROS, start=4):
    col = COL_PART[membro]
    ws.cell(i, 1, membro)
    ws.cell(i, 2, f'=SUMIF(Gastos!$G$2:$G${LAST},A{i},Gastos!$F$2:$F${LAST})')
    ws.cell(i, 3, (
        f'=SUMPRODUCT((Gastos!$D$2:$D${LAST}="Compartilhado")'
        f'*(Gastos!{col}$2:{col}${LAST}=TRUE)'
        f'*(Gastos!$F$2:$F${LAST}/Gastos!${div_col}$2:${div_col}${LAST}))'
    ))
    ws.cell(i, 4, (
        f'=SUMPRODUCT((Gastos!$D$2:$D${LAST}="Pessoal")'
        f'*(Gastos!$G$2:$G${LAST}=A{i})*Gastos!$F$2:$F${LAST})'
    ))
    ws.cell(i, 5, f"=C{i}+D{i}")
    ws.cell(i, 6, (
        f'=SUMPRODUCT((Gastos!$D$2:$D${LAST}="Compartilhado")'
        f'*(Gastos!$G$2:$G${LAST}=A{i})*Gastos!$F$2:$F${LAST})-C{i}'
    ))

ws.conditional_formatting.add(f"F4:F{3 + len(MEMBROS)}", CellIsRule(operator="greaterThan", formula=["0"], fill=pos_fill))
ws.conditional_formatting.add(f"F4:F{3 + len(MEMBROS)}", CellIsRule(operator="lessThan", formula=["0"], fill=neg_fill))
ws.append([])
ws.cell(3 + len(MEMBROS) + 2, 1, "Saldo positivo = a receber dos outros · negativo = deve ao grupo")
autosize(ws)

# ---------- Acertos ----------
ws = wb.create_sheet("Acertos")
ws.append(["Data", "De", "Para", "Valor", "Moeda", "Forma", "Obs"])
style_header(ws, 7)
dv_de = DataValidation(type="list", formula1='"%s"' % ",".join(MEMBROS), allow_blank=True)
dv_para = DataValidation(type="list", formula1='"%s"' % ",".join(MEMBROS), allow_blank=True)
dv_forma = DataValidation(type="list", formula1='"Pix,Zelle,Dinheiro,Outro"', allow_blank=True)
for dv in (dv_de, dv_para, dv_forma):
    ws.add_data_validation(dv)
for r in range(2, 42):
    dv_de.add(f"B{r}")
    dv_para.add(f"C{r}")
    dv_forma.add(f"F{r}")
ws.freeze_panes = "A2"
autosize(ws)

# ---------- Categorias ----------
ws = wb.create_sheet("Categorias")
ws.append(["Categoria", "Tipo sugerido", "Obs"])
style_header(ws, 3)
cats = [
    ("Casa / Airbnb", "Compartilhado", "Dividir entre quem usa"),
    ("Carro / combustivel", "Compartilhado", ""),
    ("Mercado", "Compartilhado", ""),
    ("Ingressos parque", "Pessoal", "Quem comprou paga"),
    ("Refeicao parque", "Pessoal", ""),
    ("eSIM / internet", "Pessoal", ""),
    ("Seguro viagem", "Pessoal", ""),
    ("Compras / outlet", "Pessoal", ""),
    ("Outro compartilhado", "Compartilhado", ""),
    ("Outro pessoal", "Pessoal", ""),
]
for row in cats:
    ws.append(list(row))
autosize(ws)

out = Path(__file__).resolve().parent / "gastos-compartilhados.xlsx"
wb.save(out)
print(f"Gerado: {out}")
print("Membros:", ", ".join(MEMBROS))
